import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ROW_FILL_ALT = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
THIN_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _apply_styles(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = ROW_FILL_ALT if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                col_letter = cell.column_letter
                content_len = len(str(cell.value))
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), content_len)

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = max(12, min(width + 2, 50))


def export_sevk_ozet(sevkler, buf):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sevk Ozeti"

    headers = [
        "Sevk No", "Tarih", "Mağaza/Alıcı", "Ürünler",
        "Nakliye (TL)", "İşçilik (TL)", "KDV %", "Toplam (TL)", "Durum"
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, sevk in enumerate(sevkler, 2):
        if sevk.magaza:
            magaza_adi = sevk.magaza.ad + "/" + sevk.magaza.sehir.ad
        else:
            magaza_adi = sevk.alici_adi or "-"

        urunler = ", ".join([f"{k.urun.ad} x{int(k.miktar)}" for k in sevk.kalemler])

        ara = (sevk.nakliye_ucreti or 0) + (sevk.iscilik or 0) + sum(g.tutar for g in sevk.giderler)
        kdv = ara * (sevk.kdv_oran or 0) / 100
        toplam = ara + kdv

        row_data = [
            sevk.id,
            sevk.tarih,
            magaza_adi,
            urunler,
            sevk.nakliye_ucreti or 0,
            sevk.iscilik or 0,
            sevk.kdv_oran or 0,
            round(toplam, 2),
            sevk.teslim_durumu,
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _apply_styles(ws, headers)
    wb.save(buf)


def export_magaza_maliyet(magazalar, buf):
    wb = Workbook()
    ws = wb.active
    ws.title = "Magaza Maliyet"

    headers = [
        "Mağaza", "Şehir", "Sevk Sayısı",
        "Nakliye Toplam (TL)", "İşçilik Toplam (TL)", "Genel Gider (TL)", "Toplam (TL)"
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, magaza in enumerate(magazalar, 2):
        sevk_sayisi = len(magaza.sevkler)
        nakliye_toplam = sum(s.nakliye_ucreti or 0 for s in magaza.sevkler)
        iscilik_toplam = sum(s.iscilik or 0 for s in magaza.sevkler)
        genel_gider = sum(g.tutar for s in magaza.sevkler for g in s.giderler)
        toplam = nakliye_toplam + iscilik_toplam + genel_gider

        row_data = [
            magaza.ad,
            magaza.sehir.ad,
            sevk_sayisi,
            round(nakliye_toplam, 2),
            round(iscilik_toplam, 2),
            round(genel_gider, 2),
            round(toplam, 2),
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _apply_styles(ws, headers)
    wb.save(buf)


def export_stok(ozet, buf):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Durumu"

    headers = [
        "Ürün Kodu", "Ürün Adı", "Birim",
        "Depo Bakiyesi", "Rezerve", "Kullanılabilir"
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, item in enumerate(ozet, 2):
        urun = item["urun"]
        row_data = [
            urun.kod,
            urun.ad,
            urun.birim,
            item["bakiye"],
            item["rezerve"],
            item["kullanilabilir"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _apply_styles(ws, headers)
    wb.save(buf)


def export_ssh(bildirimleri, buf):
    wb = Workbook()
    ws = wb.active
    ws.title = "SSH Bildirimleri"

    headers = [
        "Tarih", "Mağaza", "Ürün", "Paket No",
        "Hasar Açıklaması", "Talep Adet", "Durum", "Admin Notu"
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, bildirim in enumerate(bildirimleri, 2):
        paket_adi = bildirim.paket.paket_adi if bildirim.paket else "-"
        row_data = [
            bildirim.tarih,
            bildirim.magaza.ad if bildirim.magaza else "-",
            bildirim.urun.ad if bildirim.urun else "-",
            paket_adi,
            bildirim.hasar_aciklamasi,
            bildirim.talep_miktar,
            bildirim.durum,
            bildirim.admin_notu or "",
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _apply_styles(ws, headers)
    wb.save(buf)
