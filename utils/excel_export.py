import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

H_FILL = PatternFill("solid", fgColor="1e293b")
H_FONT = Font(color="FFFFFF", bold=True)


def _header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 45)


def export_sevk_ozet(sevkler, buf):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sevk Ozeti"
    _header(ws, ["ID", "Tarih", "Sehir", "Magaza", "Nakliye (TL)", "Iscilik (TL)", "Gider Toplam (TL)"])
    for s in sevkler:
        try:
            sehir = s.magaza.sehir.ad if s.magaza and s.magaza.sehir else "-"
            magaza = s.magaza.ad if s.magaza else (s.alici_adi or "Serbest")
            gider = sum(g.tutar for g in s.giderler)
            ws.append([s.id, s.tarih, sehir, magaza,
                       round(s.nakliye_ucreti or 0, 2), round(s.iscilik or 0, 2), round(gider, 2)])
        except Exception:
            ws.append([s.id, s.tarih, "-", "-", 0, 0, 0])
    _autofit(ws)
    wb.save(buf)


def export_magaza_maliyet(magazalar, buf):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Magaza Maliyet"
    _header(ws, ["Sehir", "Magaza", "Sevk Sayisi", "Toplam Nakliye (TL)", "Toplam Iscilik (TL)", "Toplam Gider (TL)"])
    for m in magazalar:
        try:
            sehir = m.sehir.ad if m.sehir else "-"
            nakliye = sum((s.nakliye_ucreti or 0) for s in m.sevkler)
            iscilik = sum((s.iscilik or 0) for s in m.sevkler)
            gider = sum(g.tutar for s in m.sevkler for g in s.giderler)
            ws.append([sehir, m.ad, len(m.sevkler),
                       round(nakliye, 2), round(iscilik, 2), round(gider, 2)])
        except Exception:
            ws.append(["-", m.ad if m else "-", 0, 0, 0, 0])
    _autofit(ws)
    wb.save(buf)


def export_stok(ozet, buf):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stok Durumu"
    _header(ws, ["Kod", "Urun Adi", "Birim", "Bakiye"])
    for item in ozet:
        try:
            u = item.get("urun") or item.get("urun_obj")
            if u:
                ws.append([u.kod or "-", u.ad or "-", u.birim or "-", round(item.get("bakiye", 0), 2)])
            else:
                ws.append(["-", "-", "-", round(item.get("bakiye", 0), 2)])
        except Exception:
            ws.append(["-", "-", "-", 0])
    _autofit(ws)
    wb.save(buf)


def export_ssh(bildirimleri, buf):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SSH Bildirimleri"
    _header(ws, ["ID", "Tarih", "Magaza", "Urun", "Paket", "Hasar Aciklamasi", "Talep Miktar", "Durum"])
    for b in bildirimleri:
        try:
            magaza = b.magaza.ad if b.magaza else "-"
            urun = b.urun.ad if b.urun else "-"
            paket = b.paket.paket_adi if b.paket else "-"
            ws.append([b.id, b.tarih, magaza, urun, paket,
                       b.hasar_aciklamasi or "-", b.talep_miktar or 0, b.durum or "-"])
        except Exception:
            ws.append([b.id if b else "-", "-", "-", "-", "-", "-", 0, "-"])
    _autofit(ws)
    wb.save(buf)
